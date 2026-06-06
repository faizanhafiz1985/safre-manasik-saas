"""
Safre Manasik – QA Test Results Excel Generator
Reads qa/test_results.json and produces a polished Excel workbook.
Usage: python qa/generate_results_excel.py
"""
import json, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_GREEN  = "1B4B35"
MID_GREEN   = "2E9E6B"
LIGHT_GREEN = "E6F4EE"
YELLOW      = "C9A227"
YELLOW_LIGHT= "FFF8E1"
RED         = "C0392B"
RED_LIGHT   = "FDECEC"
GREY_HEADER = "F3F8F5"
GREY_ROW    = "FAFAFA"
WHITE       = "FFFFFF"
BORDER_CLR  = "CCCCCC"

PASS_FILL   = PatternFill("solid", fgColor=LIGHT_GREEN)
FAIL_FILL   = PatternFill("solid", fgColor=RED_LIGHT)
SKIP_FILL   = PatternFill("solid", fgColor="FFF3E0")
HEAD_FILL   = PatternFill("solid", fgColor=DARK_GREEN)
SUB_FILL    = PatternFill("solid", fgColor=MID_GREEN)
GREY_FILL   = PatternFill("solid", fgColor=GREY_HEADER)
ALT_FILL    = PatternFill("solid", fgColor=GREY_ROW)

def thin_border(all_sides=False):
    s = Side(style='thin', color=BORDER_CLR)
    if all_sides:
        return Border(left=s, right=s, top=s, bottom=s)
    return Border(bottom=s)

def header_font(sz=11, bold=True, color=WHITE):
    return Font(name='Arial', size=sz, bold=bold, color=color)

def body_font(sz=10, bold=False, color="000000"):
    return Font(name='Arial', size=sz, bold=bold, color=color)

def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ── Load results ─────────────────────────────────────────────────────────────
results_path = os.path.join(os.path.dirname(__file__), 'test_results.json')
with open(results_path) as f:
    data = json.load(f)

results = data['results']
meta    = data['meta']
run_at  = meta.get('runAt', datetime.utcnow().isoformat())

pass_n  = meta['pass']
fail_n  = meta['fail']
skip_n  = meta['skip']
total_n = meta['total']

# Separate by category
functional = [r for r in results if r['category'] == 'Functional']
negative   = [r for r in results if r['category'] == 'Negative']
technical  = [r for r in results if r['category'] == 'Technical']

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 8  # top margin

# Title block
ws1.merge_cells("A2:G2")
ws1["A2"] = "SAFRE MANASIK – QA TEST RESULTS"
ws1["A2"].font = Font(name='Arial', size=18, bold=True, color=WHITE)
ws1["A2"].fill = PatternFill("solid", fgColor=DARK_GREEN)
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 36

ws1.merge_cells("A3:G3")
ws1["A3"] = f"Test Run: {run_at[:16].replace('T',' ')} UTC"
ws1["A3"].font = Font(name='Arial', size=10, color=WHITE)
ws1["A3"].fill = PatternFill("solid", fgColor=MID_GREEN)
ws1["A3"].alignment = center()
ws1.row_dimensions[3].height = 20

ws1.row_dimensions[4].height = 12  # spacer

# Stat cards: row 5 (label) + row 6 (value)
card_data = [
    ("A", "TOTAL",   str(total_n), DARK_GREEN),
    ("B", "PASS ✓",  str(pass_n),  MID_GREEN),
    ("C", "FAIL ✗",  str(fail_n),  RED if fail_n else MID_GREEN),
    ("D", "SKIP ⊘",  str(skip_n),  YELLOW),
    ("E", "PASS RATE", f"{int(pass_n/(pass_n+fail_n)*100) if (pass_n+fail_n) else 0}%", DARK_GREEN),
]
for col_letter, lbl, val, color in card_data:
    ws1[f"{col_letter}5"] = lbl
    ws1[f"{col_letter}5"].font = Font(name='Arial', size=9, bold=True, color=WHITE)
    ws1[f"{col_letter}5"].fill = PatternFill("solid", fgColor=color)
    ws1[f"{col_letter}5"].alignment = center()
    ws1.row_dimensions[5].height = 18

    ws1[f"{col_letter}6"] = val
    ws1[f"{col_letter}6"].font = Font(name='Arial', size=22, bold=True, color=color)
    ws1[f"{col_letter}6"].fill = PatternFill("solid", fgColor=GREY_ROW)
    ws1[f"{col_letter}6"].alignment = center()
    ws1.row_dimensions[6].height = 40

ws1.row_dimensions[7].height = 14  # spacer

# Category breakdown table
ws1["A8"] = "Category"
ws1["B8"] = "Total"
ws1["C8"] = "Pass"
ws1["D8"] = "Fail"
ws1["E8"] = "Skip"
ws1["F8"] = "Pass Rate"
for col in "ABCDEF":
    c = ws1[f"{col}8"]
    c.font = header_font(10)
    c.fill = PatternFill("solid", fgColor=DARK_GREEN)
    c.alignment = center()
    c.border = thin_border(True)
ws1.row_dimensions[8].height = 20

cat_rows = [
    ("Functional",  functional),
    ("Negative",    negative),
    ("Technical",   technical),
    ("ALL",         results),
]
for i, (cat, rows) in enumerate(cat_rows, 9):
    p = sum(1 for r in rows if r['status'] == 'PASS')
    f = sum(1 for r in rows if r['status'] == 'FAIL')
    s = sum(1 for r in rows if r['status'] == 'SKIP')
    t = len(rows)
    rate = f"{int(p/(p+f)*100) if (p+f) else 0}%"
    vals = [cat, t, p, f, s, rate]
    for j, v in enumerate(vals, 1):
        c = ws1.cell(row=i, column=j, value=v)
        c.font = body_font(10, bold=(cat == 'ALL'))
        c.fill = GREY_FILL if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        if cat == 'ALL':
            c.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        c.alignment = center()
        c.border = thin_border(True)
    ws1.row_dimensions[i].height = 18

ws1.row_dimensions[13].height = 14  # spacer

# Failed tests list
ws1["A14"] = "⚠ Failed Test Cases"
ws1["A14"].font = Font(name='Arial', size=11, bold=True, color=RED)
ws1.row_dimensions[14].height = 18

failed = [r for r in results if r['status'] == 'FAIL']
if not failed:
    ws1.merge_cells("A15:G15")
    ws1["A15"] = "✓ No failed tests — all executed tests passed!"
    ws1["A15"].font = Font(name='Arial', size=10, color=MID_GREEN, bold=True)
    ws1["A15"].alignment = left(False)
else:
    headers = ["ID", "Category", "Test Name", "Expected", "Actual"]
    for j, h in enumerate(headers, 1):
        c = ws1.cell(row=15, column=j, value=h)
        c.font = header_font(9)
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = center()
    ws1.row_dimensions[15].height = 16
    for i, r in enumerate(failed, 16):
        vals = [r['id'], r['category'], r['name'], r['expected'], r['actual']]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=i, column=j, value=v)
            c.font = body_font(9)
            c.fill = FAIL_FILL
            c.alignment = left()
            c.border = thin_border()
        ws1.row_dimensions[i].height = 16

set_col_widths(ws1, [18, 16, 42, 18, 22, 12, 10])
freeze(ws1, "A5")

# ════════════════════════════════════════════════════════════════════════════
# HELPER: build a results sheet
# ════════════════════════════════════════════════════════════════════════════
STATUS_FILL = {'PASS': PASS_FILL, 'FAIL': FAIL_FILL, 'SKIP': SKIP_FILL}
STATUS_COLOR = {'PASS': MID_GREEN, 'FAIL': RED, 'SKIP': YELLOW}

def build_results_sheet(ws, title, rows, header_color):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8

    # Sheet title
    ws.merge_cells("A2:G2")
    ws["A2"] = title
    ws["A2"].font = Font(name='Arial', size=14, bold=True, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=header_color)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 8  # spacer

    # Column headers
    headers = ["Test ID", "Category", "Test Name", "Status", "Expected", "Actual", "Notes"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = header_font(10)
        c.fill = PatternFill("solid", fgColor=header_color)
        c.alignment = center()
        c.border = thin_border(True)
    ws.row_dimensions[4].height = 20
    freeze(ws, "A5")

    # Data rows
    for i, r in enumerate(rows, 5):
        status = r.get('status', 'SKIP')
        fill   = STATUS_FILL.get(status, ALT_FILL)
        vals   = [r['id'], r['category'], r['name'], status, r['expected'], r['actual'], r.get('notes','')]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = fill if status != 'PASS' else (PASS_FILL if i % 2 == 0 else PatternFill("solid", fgColor=WHITE))
            c.font = body_font(9, bold=(j == 4), color=STATUS_COLOR.get(status, "000000") if j == 4 else "000000")
            c.alignment = left(j in (3, 6, 7))
            c.border = thin_border()
        ws.row_dimensions[i].height = 16

    set_col_widths(ws, [10, 12, 40, 9, 18, 28, 22])

# ════════════════════════════════════════════════════════════════════════════
# SHEETS 2–4: per-category results
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Functional Tests")
build_results_sheet(ws2, "Functional Tests", functional, DARK_GREEN)

ws3 = wb.create_sheet("Negative Tests")
build_results_sheet(ws3, "Negative Tests", negative, "8B0000")

ws4 = wb.create_sheet("Technical Tests")
build_results_sheet(ws4, "Technical Tests", technical, "1A237E")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 – ALL RESULTS (flat list)
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("All Results")
build_results_sheet(ws5, "All 65 Test Cases – Combined Results", results, DARK_GREEN)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 – DEFECT LOG
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Defect Log")
ws6.sheet_view.showGridLines = False
ws6.row_dimensions[1].height = 8

ws6.merge_cells("A2:H2")
ws6["A2"] = "DEFECT LOG"
ws6["A2"].font = Font(name='Arial', size=14, bold=True, color=WHITE)
ws6["A2"].fill = PatternFill("solid", fgColor=RED)
ws6["A2"].alignment = center()
ws6.row_dimensions[2].height = 26
ws6.row_dimensions[3].height = 8

defect_headers = ["Defect #", "Test ID", "Category", "Description", "Severity", "Status", "Fix Action", "Fixed In"]
for j, h in enumerate(defect_headers, 1):
    c = ws6.cell(row=4, column=j, value=h)
    c.font = header_font(10)
    c.fill = PatternFill("solid", fgColor=RED)
    c.alignment = center()
    c.border = thin_border(True)
ws6.row_dimensions[4].height = 20
freeze(ws6, "A5")

failed_for_defect = [r for r in results if r['status'] == 'FAIL']
for i, r in enumerate(failed_for_defect, 5):
    severity = "HIGH" if r['category'] == 'Functional' else "MEDIUM"
    vals = [f"DEF-{i-4:03d}", r['id'], r['category'],
            f"{r['name']} — expected {r['expected']}, got {r['actual']}",
            severity, "OPEN", "", ""]
    for j, v in enumerate(vals, 1):
        c = ws6.cell(row=i, column=j, value=v)
        c.fill = FAIL_FILL
        c.font = body_font(9)
        c.alignment = left(j == 4)
        c.border = thin_border()
    ws6.row_dimensions[i].height = 18

if not failed_for_defect:
    ws6.merge_cells("A5:H5")
    ws6["A5"] = "✓ No defects found. All executed tests passed!"
    ws6["A5"].font = Font(name='Arial', size=11, bold=True, color=MID_GREEN)
    ws6["A5"].alignment = left(False)

set_col_widths(ws6, [10, 10, 12, 48, 10, 10, 28, 12])

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out_path = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - Balubaid Group of Companies",
    "Desktop",
    "safre-manasik-qa-results.xlsx"
)
# Fallback to current dir if OneDrive path doesn't exist
if not os.path.isdir(os.path.dirname(out_path)):
    out_path = os.path.join(os.path.dirname(__file__), "safre-manasik-qa-results.xlsx")

wb.save(out_path)
print(f"DONE: Results workbook saved: {out_path}")
print(f"  TOTAL={total_n}  PASS={pass_n}  FAIL={fail_n}  SKIP={skip_n}")
